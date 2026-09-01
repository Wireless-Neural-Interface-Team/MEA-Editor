"""
I/O layer for the electrode array editor.

    Native format (Save / Open):
    mea_editor JSON — electrodes, pads, orientation markers, extra attributes,
    shapes, map labels, per-item label position. No probeinterface dependency.

Legacy inputs still accepted:
    1) mea_editor JSON (current)
    2) older custom editor JSON (channel_index / contact_id)
    3) probeinterface JSON (migrated into native fields)

SpikeInterface export:
    probeinterface JSON for recording contacts only (pads are not SI contacts):
    - device_channel_indices = channel ID derived from INTAN ID
    - contact_ids = Contact ID from Manufacturer ID (fallback: INTAN ID)
    - contact_annotations = native IDs (including shank_id), extras, and the
      first linked pad (pad_id plus geometry)
    - probe annotations = electrode attribute schema and visible map labels
      so a file exported from this editor can be reopened with less data loss

XLSX exports:
    analysis table (channel / row / col) and a full array workbook
    (electrodes sheet + pads sheet + orientation-markers sheet + attribute
    schema). Geometry uses radius / width / height (circle: radius; square:
    width and height; rect: independent extents). Extra electrode attributes
    follow the file schema on both electrode and pad sheets. Orientation
    markers are Excel-only (not SpikeInterface contacts).
"""

from __future__ import annotations

import json
import re
from collections import Counter
from collections.abc import Iterable
from dataclasses import dataclass, field
from typing import Any

from ._version import __version__ as EDITOR_VERSION
from .attribute_schema import (
    AttributeSpec,
    coerce_value,
    extra_specs,
    fill_electrodes_extras,
    schema_from_payload,
    schema_to_payload,
)
from .contact_shape import (
    DEFAULT_ELECTRODE_SHAPE,
    DEFAULT_PAD_SHAPE,
    export_contact_sizes,
    extents_from_probe_params,
    normalize_contact_shape,
    probe_shape_params,
)
from .electrode import (
    BUILTIN_ATTRIBUTE_KEYS,
    DEFAULT_INTAN_ID,
    DEFAULT_LABEL_ORIENTATION,
    DEFAULT_LABEL_POSITION,
    DEFAULT_MANUFACTURER_ID,
    DEFAULT_MAP_LABEL_KEYS,
    DEFAULT_PLANE_AXIS,
    DEFAULT_RADIUS,
    DEFAULT_SHAPE,
    Electrode,
    normalize_label_orientation,
    normalize_label_position,
)
from .orientation_marker import (
    DEFAULT_MARKER_RADIUS,
    DEFAULT_MARKER_SHAPE,
    OrientationMarker,
)
from .pad import (
    DEFAULT_PAD_RADIUS,
    Pad,
)

NATIVE_SPECIFICATION = "mea_editor"
NATIVE_VERSION = "1.11"
MIN_RADIUS = 0.001
DEFAULT_UNITS = "um"
PROBEINTERFACE_SPEC = "probeinterface"
PROBEINTERFACE_VERSION = "0.3.1"
INTAN_CHANNELS_PER_PORT = 32
SI_CHANNEL_ID_KEY = "channel ID"
SI_CONTACT_ID_KEY = "Contact ID"
MEA_EDITOR_SCHEMA_ANNOTATION_KEY = "mea_editor_electrode_attributes"
MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY = "mea_editor_map_labels"
_MISSING = object()
PAD_GEOMETRY_ANNOTATION_KEYS = ("pad_id", "pad_x", "pad_y", "pad_shape", "pad_radius", "pad_height")
NATIVE_CONTACT_ANNOTATION_KEYS = frozenset(
    {
        SI_CHANNEL_ID_KEY,
        SI_CONTACT_ID_KEY,
        "channel_id",
        "contact_id",
        "potentiostat_id",
        "intan_id",
        "manufacturer_id",
        "shank_id",
        "pad_interface_id",
        "pad_system_id",
        "pad_id",
        "pad_x",
        "pad_y",
        "pad_shape",
        "pad_radius",
        "pad_height",
        "eid",
        MEA_EDITOR_SCHEMA_ANNOTATION_KEY,
        MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY,
    }
)

_INTAN_PORT_RE = re.compile(r"^([A-Z])[\-_]?(\d+)$", re.IGNORECASE)
_INTAN_NC_RE = re.compile(r"^NC(\d*)$", re.IGNORECASE)

_ELECTRODE_KNOWN_KEYS = frozenset(
    {
        "eid",
        "x",
        "y",
        "radius",
        "height",
        "enabled",  # dropped field; ignored on load so it does not become an extra
        "potentiostat_id",
        "intan_id",
        "manufacturer_id",
        "contact_plane_axis",
        "shank_id",
        "shape",
        "label_position",
        "label_orientation",
        "channel_index",
        "contact_id",
        "attributes",
        "extra",
    }
)


@dataclass
class ArrayDocument:
    """Loaded array: models, schema, units, and visible map-label keys."""

    electrodes: list[Electrode]
    pads: list[Pad] = field(default_factory=list)
    orientation_markers: list[OrientationMarker] = field(default_factory=list)
    si_units: str = DEFAULT_UNITS
    electrode_attributes: list[AttributeSpec] = field(default_factory=list)
    map_labels: list[str] = field(default_factory=list)


def normalize_map_labels(raw: Any, schema: list[AttributeSpec]) -> list[str]:
    """
    Visible map-label keys in schema order.

    Missing / invalid values fall back to the editor defaults. An explicit
    empty list means every label is hidden.
    """
    valid = [spec.key for spec in schema]
    valid_set = set(valid)
    if raw is _MISSING or raw is None:
        return [key for key in DEFAULT_MAP_LABEL_KEYS if key in valid_set]
    if not isinstance(raw, list):
        return [key for key in DEFAULT_MAP_LABEL_KEYS if key in valid_set]
    selected = {str(item).strip() for item in raw if str(item).strip()}
    return [key for key in valid if key in selected]


def _parse_contact_plane_axis(raw_value: Any) -> tuple[float, float, float, float]:
    """
    Parse contact plane axis from generic input (list/tuple).

    Expected order: (plane_axis_x_0, plane_axis_x_1, plane_axis_y_0, plane_axis_y_1).
    """
    if not isinstance(raw_value, (list, tuple)) or len(raw_value) != 4:
        return DEFAULT_PLANE_AXIS
    try:
        x0, x1, y0, y1 = (float(v) for v in raw_value)
    except (TypeError, ValueError):
        return DEFAULT_PLANE_AXIS
    return (x0, x1, y0, y1)


def _as_float(value: Any, default: float) -> float:
    try:
        if value is None:
            return default
        number = float(value)
    except (TypeError, ValueError):
        return default
    if number != number:  # NaN
        return default
    return number


def _as_int(value: Any, default: int) -> int:
    try:
        if value is None or isinstance(value, bool):
            return default
        return int(value)
    except (TypeError, ValueError):
        return default


def _as_str(value: Any, default: str) -> str:
    if value is None:
        return default
    text = str(value).strip() if isinstance(value, str) else str(value)
    return text if text != "" or default == "" else default


def si_units_for_probeinterface(si_units: str) -> str:
    """Map editor units onto the probeinterface enum (`um` or `mm`)."""
    text = (si_units or DEFAULT_UNITS).strip()
    lowered = text.lower()
    if lowered in {"um", "micron", "microns", "micrometer", "micrometre"} or text in {"µm", "μm"}:
        return "um"
    if lowered in {"mm", "millimeter", "millimetre", "millimeters", "millimetres"}:
        return "mm"
    return "um"


def _annotation_at(annotations: Any, key: str, index: int) -> Any:
    if not isinstance(annotations, dict):
        return None
    values = annotations.get(key)
    if isinstance(values, list) and index < len(values):
        return values[index]
    return None


def _first_pad_by_electrode(pads: list[Pad] | None) -> dict[int, Pad]:
    """Map each electrode eid to its first pad (lowest pad_id). Later pads are ignored."""
    mapping: dict[int, Pad] = {}
    for pad in sorted(pads or [], key=lambda item: item.pad_id):
        mapping.setdefault(pad.electrode_eid, pad)
    return mapping


def _contact_plane_axis_text(axis: tuple[float, float, float, float]) -> str:
    return ", ".join(f"{value:g}" for value in axis)


def _extra_values(model: Electrode | None, extras: list[AttributeSpec]) -> list[Any]:
    """Row values for extra schema attributes; empty cells when there is no electrode."""
    if model is None:
        return ["" for _ in extras]
    return [
        coerce_value(spec.value_type, model.get_attribute(spec.key), spec.default)
        for spec in extras
    ]


def _schema_payload_from_probe(probe: dict[str, Any]) -> Any:
    """Read the native attribute schema stored on a probeinterface probe."""
    annotations = probe.get("annotations")
    if not isinstance(annotations, dict):
        return None
    return annotations.get(MEA_EDITOR_SCHEMA_ANNOTATION_KEY)


def _map_labels_from_probe(probe: dict[str, Any]) -> Any:
    """Read visible map-label keys stored on a probeinterface probe."""
    annotations = probe.get("annotations")
    if not isinstance(annotations, dict):
        return _MISSING
    if MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY not in annotations:
        return _MISSING
    return annotations.get(MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY)


def format_intan_id(index: int, channels_per_port: int = INTAN_CHANNELS_PER_PORT) -> str:
    """
    Build a unique INTAN ID from a sequential index.

    0 -> A-000, 31 -> A-031, 32 -> B-000. Indexes beyond Z wrap to a plain integer.
    """
    if index < 0:
        raise ValueError("INTAN index must be >= 0.")
    port = index // channels_per_port
    channel = index % channels_per_port
    if port >= 26:
        return str(index)
    return f"{chr(ord('A') + port)}-{channel:03d}"


def intan_id_to_channel_id(intan_id: str, channels_per_port: int = INTAN_CHANNELS_PER_PORT) -> int:
    """
    Convert an INTAN ID into a probeinterface / SpikeInterface channel ID.

    Accepted forms:
    - integer string: "12" -> 12
    - port-channel: "A-003", "A003", "D-018" -> port * channels_per_port + channel
      (A=0 … Z=25; channel must be 0 … channels_per_port-1, 32 by default)
    - disconnected: "NC", "NC1", "NC3" -> -1

    Raises:
        ValueError: if the INTAN ID cannot be converted.
    """
    text = str(intan_id).strip()
    if not text:
        raise ValueError("INTAN ID is empty.")

    if text.lstrip("-").isdigit():
        return int(text)

    nc_match = _INTAN_NC_RE.fullmatch(text)
    if nc_match:
        return -1

    port_match = _INTAN_PORT_RE.fullmatch(text)
    if port_match:
        port = ord(port_match.group(1).upper()) - ord("A")
        channel = int(port_match.group(2))
        if channel < 0 or channel >= channels_per_port:
            raise ValueError(
                f"INTAN channel in « {intan_id} » must be 0–{channels_per_port - 1} "
                f"(A-032 is B-000, not a valid A-port channel)."
            )
        return port * channels_per_port + channel

    raise ValueError(
        f"Cannot convert INTAN ID « {intan_id} » to a channel ID. "
        "Use an integer, a port-channel like A-003, or NC for disconnected."
    )


def try_intan_channel_id(intan_id: str, channels_per_port: int = INTAN_CHANNELS_PER_PORT) -> int | None:
    """Return the SpikeInterface channel ID, or None if the INTAN ID is invalid."""
    try:
        return intan_id_to_channel_id(intan_id, channels_per_port)
    except ValueError:
        return None


def _extra_from_electrode_dict(el: dict[str, Any]) -> dict[str, Any]:
    """Collect file-defined extra attributes from nested or inline keys."""
    extra: dict[str, Any] = {}
    nested = el.get("attributes", el.get("extra"))
    if isinstance(nested, dict):
        for key, value in nested.items():
            key_text = str(key).strip()
            if key_text and key_text not in _ELECTRODE_KNOWN_KEYS and key_text not in BUILTIN_ATTRIBUTE_KEYS:
                extra[key_text] = value
    for key, value in el.items():
        key_text = str(key).strip()
        if key_text in _ELECTRODE_KNOWN_KEYS or key_text in BUILTIN_ATTRIBUTE_KEYS:
            continue
        extra[key_text] = value
    return extra


def _electrode_from_native_dict(el: dict[str, Any], fallback_index: int) -> Electrode:
    """Build an Electrode from a native or legacy custom JSON object."""
    eid = _as_int(el.get("eid"), fallback_index)
    potentiostat_id = el.get("potentiostat_id", el.get("channel_index", fallback_index))
    intan_id = el.get("intan_id", el.get("contact_id", DEFAULT_INTAN_ID))
    manufacturer_id = el.get("manufacturer_id", DEFAULT_MANUFACTURER_ID)
    radius = max(_as_float(el.get("radius"), DEFAULT_RADIUS), MIN_RADIUS)
    height = max(_as_float(el.get("height"), 0.0), 0.0)
    shape = normalize_contact_shape(_as_str(el.get("shape"), DEFAULT_SHAPE), DEFAULT_SHAPE)
    return Electrode(
        eid=eid,
        x=_as_float(el.get("x"), 0.0),
        y=_as_float(el.get("y"), 0.0),
        radius=radius,
        height=height,
        potentiostat_id=_as_int(potentiostat_id, fallback_index),
        intan_id=_as_str(intan_id, DEFAULT_INTAN_ID),
        manufacturer_id=_as_str(manufacturer_id, DEFAULT_MANUFACTURER_ID),
        contact_plane_axis=_parse_contact_plane_axis(el.get("contact_plane_axis")),
        shank_id=_as_str(el.get("shank_id"), ""),
        shape=shape,
        label_position=normalize_label_position(el.get("label_position", DEFAULT_LABEL_POSITION)),
        label_orientation=normalize_label_orientation(
            el.get("label_orientation", DEFAULT_LABEL_ORIENTATION)
        ),
        extra=_extra_from_electrode_dict(el),
    )


def _load_native_or_legacy_custom(
    data: dict[str, Any],
) -> tuple[list[Electrode], str, list[AttributeSpec]]:
    raw_list = data.get("electrodes")
    if not isinstance(raw_list, list):
        raise ValueError("No electrodes found in file.")
    si_units = str(data.get("si_units", DEFAULT_UNITS) or DEFAULT_UNITS)
    models: list[Electrode] = []
    for i, el in enumerate(raw_list):
        if not isinstance(el, dict):
            continue
        models.append(_electrode_from_native_dict(el, i))
    if not models:
        raise ValueError("No electrodes found in file.")
    schema = schema_from_payload(data.get("electrode_attributes"), models)
    fill_electrodes_extras(models, schema, prune=False)
    return models, si_units, schema


def _plane_axis_from_probe_entry(axes_entry: Any) -> tuple[float, float, float, float]:
    """
    probeinterface stores plane axes as [[x0, x1], [y0, y1]].
    """
    if not isinstance(axes_entry, (list, tuple)) or len(axes_entry) != 2:
        return DEFAULT_PLANE_AXIS
    row_x, row_y = axes_entry
    if not isinstance(row_x, (list, tuple)) or len(row_x) != 2:
        return DEFAULT_PLANE_AXIS
    if not isinstance(row_y, (list, tuple)) or len(row_y) != 2:
        return DEFAULT_PLANE_AXIS
    try:
        return (float(row_x[0]), float(row_x[1]), float(row_y[0]), float(row_y[1]))
    except (TypeError, ValueError):
        return DEFAULT_PLANE_AXIS


def _extents_from_probe_shape(shape: str, params: Any) -> tuple[float, float]:
    return extents_from_probe_params(shape, params, DEFAULT_RADIUS, MIN_RADIUS)


def _pad_from_contact_annotations(
    annotations: Any,
    index: int,
    electrode_eid: int,
    fallback_pad_id: int,
) -> Pad | None:
    """Rebuild a pad from SpikeInterface contact annotations when geometry is present."""
    x_raw = _annotation_at(annotations, "pad_x", index)
    y_raw = _annotation_at(annotations, "pad_y", index)
    if x_raw in (None, "") or y_raw in (None, ""):
        return None
    pad_id_raw = _annotation_at(annotations, "pad_id", index)
    pad_id = fallback_pad_id if pad_id_raw in (None, "") else _as_int(pad_id_raw, fallback_pad_id)
    shape = normalize_contact_shape(
        _as_str(_annotation_at(annotations, "pad_shape", index), DEFAULT_PAD_SHAPE),
        DEFAULT_PAD_SHAPE,
    )
    radius = max(_as_float(_annotation_at(annotations, "pad_radius", index), DEFAULT_PAD_RADIUS), MIN_RADIUS)
    height = max(_as_float(_annotation_at(annotations, "pad_height", index), 0.0), 0.0)
    return Pad(
        pad_id=pad_id,
        electrode_eid=electrode_eid,
        x=_as_float(x_raw, 0.0),
        y=_as_float(y_raw, 0.0),
        radius=radius,
        height=height,
        shape=shape,
    )


def _load_probeinterface(data: dict[str, Any]) -> tuple[list[Electrode], list[Pad], str, Any, Any]:
    """
    Load a probeinterface JSON without importing probeinterface.

    Default mapping (plain probeinterface files):
    - device_channel_indices -> potentiostat_id
    - contact_ids -> intan_id
    - manufacturer_id left empty

    Files exported by this editor also store native fields in
    `contact_annotations` and the attribute schema / map labels in probe
    annotations; those override the default mapping when present.
    Extra annotation keys become electrode extra attributes.
    Pads are restored only when pad geometry (`pad_x` / `pad_y`) is stored
    in contact_annotations (mea_editor exports).
    """
    probes = data.get("probes")
    if not isinstance(probes, list) or not probes:
        raise ValueError("No probes found in probeinterface file.")

    models: list[Electrode] = []
    pads: list[Pad] = []
    si_units = DEFAULT_UNITS
    fallback_eid = 0
    fallback_pad_id = 0
    schema_payload: Any = None
    map_labels_payload: Any = _MISSING
    for probe in probes:
        if not isinstance(probe, dict):
            continue
        si_units = str(probe.get("si_units", si_units) or si_units)
        positions = probe.get("contact_positions") or []
        if not isinstance(positions, list):
            continue
        n = len(positions)
        plane_axes = probe.get("contact_plane_axes") or []
        shapes = probe.get("contact_shapes") or []
        shape_params = probe.get("contact_shape_params") or []
        device_channels = probe.get("device_channel_indices")
        contact_ids = probe.get("contact_ids")
        shank_ids = probe.get("shank_ids")
        annotations = probe.get("contact_annotations")
        if schema_payload is None:
            schema_payload = _schema_payload_from_probe(probe)
        if map_labels_payload is _MISSING:
            map_labels_payload = _map_labels_from_probe(probe)

        for i in range(n):
            pos = positions[i]
            if not isinstance(pos, (list, tuple)) or len(pos) < 2:
                continue
            x = _as_float(pos[0], 0.0)
            y = _as_float(pos[1], 0.0)
            shape = DEFAULT_SHAPE
            if isinstance(shapes, list) and i < len(shapes):
                shape = normalize_contact_shape(
                    _as_str(shapes[i], DEFAULT_SHAPE),
                    DEFAULT_SHAPE,
                )
            params = shape_params[i] if isinstance(shape_params, list) and i < len(shape_params) else {}
            radius, height = _extents_from_probe_shape(shape, params)
            plane = DEFAULT_PLANE_AXIS
            if isinstance(plane_axes, list) and i < len(plane_axes):
                plane = _plane_axis_from_probe_entry(plane_axes[i])

            eid = fallback_eid
            eid_ann = _annotation_at(annotations, "eid", i)
            if eid_ann is not None:
                eid = _as_int(eid_ann, fallback_eid)

            potentiostat_id = eid
            if isinstance(device_channels, list) and i < len(device_channels) and device_channels[i] is not None:
                potentiostat_id = _as_int(device_channels[i], eid)
            pot_ann = _annotation_at(annotations, "potentiostat_id", i)
            if pot_ann is not None:
                potentiostat_id = _as_int(pot_ann, potentiostat_id)

            intan_id = DEFAULT_INTAN_ID
            if isinstance(contact_ids, list) and i < len(contact_ids) and contact_ids[i] is not None:
                intan_id = _as_str(contact_ids[i], DEFAULT_INTAN_ID)
            intan_ann = _annotation_at(annotations, "intan_id", i)
            if intan_ann is not None:
                intan_id = _as_str(intan_ann, intan_id)

            manufacturer_id = DEFAULT_MANUFACTURER_ID
            man_ann = _annotation_at(annotations, "manufacturer_id", i)
            if man_ann is not None:
                manufacturer_id = _as_str(man_ann, DEFAULT_MANUFACTURER_ID)

            shank_id = ""
            if isinstance(shank_ids, list) and i < len(shank_ids) and shank_ids[i] is not None:
                shank_id = _as_str(shank_ids[i], "")
            shank_ann = _annotation_at(annotations, "shank_id", i)
            if shank_ann is not None:
                shank_id = _as_str(shank_ann, shank_id)

            extra: dict[str, Any] = {}
            if isinstance(annotations, dict):
                for key, values in annotations.items():
                    key_text = str(key).strip()
                    if (
                        not key_text
                        or key_text in NATIVE_CONTACT_ANNOTATION_KEYS
                        or key_text in BUILTIN_ATTRIBUTE_KEYS
                        or key_text in _ELECTRODE_KNOWN_KEYS
                    ):
                        continue
                    extra[key_text] = _annotation_at(annotations, key_text, i)

            models.append(
                Electrode(
                    eid=eid,
                    x=x,
                    y=y,
                    radius=radius,
                    height=height,
                    potentiostat_id=potentiostat_id,
                    intan_id=intan_id,
                    manufacturer_id=manufacturer_id,
                    contact_plane_axis=plane,
                    shank_id=shank_id,
                    shape=shape,
                    extra=extra,
                )
            )
            pad = _pad_from_contact_annotations(annotations, i, eid, fallback_pad_id)
            if pad is not None:
                pads.append(pad)
                fallback_pad_id = max(fallback_pad_id, pad.pad_id) + 1
            fallback_eid = max(fallback_eid, eid) + 1

    if not models:
        raise ValueError("No contacts found in probeinterface file.")
    return models, pads, si_units, schema_payload, map_labels_payload


def _pad_from_native_dict(raw: dict[str, Any], fallback_index: int) -> Pad:
    """Build a Pad from a native JSON object.

    Older files may still contain `pid`, `interface_id`, or `system_id`;
    those keys are ignored except that `pid` is accepted as an alias for `pad_id`.
    """
    pad_id = _as_int(raw.get("pad_id", raw.get("pid")), fallback_index)
    electrode_eid = _as_int(raw.get("electrode_eid"), -1)
    radius = max(_as_float(raw.get("radius"), DEFAULT_PAD_RADIUS), MIN_RADIUS)
    height = max(_as_float(raw.get("height"), 0.0), 0.0)
    shape = normalize_contact_shape(_as_str(raw.get("shape"), DEFAULT_PAD_SHAPE), DEFAULT_PAD_SHAPE)
    return Pad(
        pad_id=pad_id,
        electrode_eid=electrode_eid,
        x=_as_float(raw.get("x"), 0.0),
        y=_as_float(raw.get("y"), 0.0),
        radius=radius,
        height=height,
        shape=shape,
        label_position=normalize_label_position(raw.get("label_position", DEFAULT_LABEL_POSITION)),
        label_orientation=normalize_label_orientation(
            raw.get("label_orientation", DEFAULT_LABEL_ORIENTATION)
        ),
    )


def _load_pads_from_payload(data: dict[str, Any]) -> list[Pad]:
    """Load pads from native JSON. Missing or invalid lists yield an empty list."""
    raw_list = data.get("pads")
    if not isinstance(raw_list, list):
        return []
    pads: list[Pad] = []
    for i, raw in enumerate(raw_list):
        if not isinstance(raw, dict):
            continue
        pads.append(_pad_from_native_dict(raw, i))
    return pads


def _orientation_marker_from_native_dict(raw: dict[str, Any], fallback_index: int) -> OrientationMarker:
    """Build an OrientationMarker from a native JSON object.

    Current files store `shape` / `radius` / `height` like pads. Older files
    stored a square `side` (full extent); that is converted to a square with
    `radius = side / 2`.
    """
    marker_id = _as_int(raw.get("marker_id"), fallback_index)
    if "radius" in raw:
        shape = normalize_contact_shape(
            _as_str(raw.get("shape"), DEFAULT_MARKER_SHAPE),
            DEFAULT_MARKER_SHAPE,
        )
        radius = max(_as_float(raw.get("radius"), DEFAULT_MARKER_RADIUS), MIN_RADIUS)
        height = max(_as_float(raw.get("height"), 0.0), 0.0)
    else:
        side = max(_as_float(raw.get("side"), DEFAULT_MARKER_RADIUS * 2.0), MIN_RADIUS)
        shape = DEFAULT_MARKER_SHAPE
        radius = max(side / 2.0, MIN_RADIUS)
        height = 0.0
    return OrientationMarker(
        marker_id=marker_id,
        x=_as_float(raw.get("x"), 0.0),
        y=_as_float(raw.get("y"), 0.0),
        radius=radius,
        height=height,
        shape=shape,
        label_position=normalize_label_position(raw.get("label_position", DEFAULT_LABEL_POSITION)),
        label_orientation=normalize_label_orientation(
            raw.get("label_orientation", DEFAULT_LABEL_ORIENTATION)
        ),
    )


def _load_orientation_markers_from_payload(data: dict[str, Any]) -> list[OrientationMarker]:
    """Load orientation markers from native JSON. Missing lists yield an empty list."""
    raw_list = data.get("orientation_markers")
    if not isinstance(raw_list, list):
        return []
    markers: list[OrientationMarker] = []
    for i, raw in enumerate(raw_list):
        if not isinstance(raw, dict):
            continue
        markers.append(_orientation_marker_from_native_dict(raw, i))
    return markers


def load_array_document(path: str) -> ArrayDocument:
    """
    Load a full array document from a JSON file.

    Supported formats:
    1) mea_editor native JSON (pads, orientation markers, map_labels,
       label_position, and label_orientation optional on older files)
    2) legacy custom editor JSON
    3) probeinterface JSON (read without the probeinterface package)

    Raises:
        ValueError: unsupported format or empty content.
        OSError / json.JSONDecodeError: file errors (propagated).
    """
    with open(path, "r", encoding="utf-8") as fh:
        data = json.load(fh)

    if not isinstance(data, dict):
        raise ValueError("Unsupported file format.")

    spec = data.get("specification")
    if spec == PROBEINTERFACE_SPEC:
        models, pads, units, schema_payload, map_labels_payload = _load_probeinterface(data)
        schema = schema_from_payload(schema_payload, models)
        fill_electrodes_extras(models, schema, prune=False)
        return ArrayDocument(
            electrodes=models,
            pads=pads,
            orientation_markers=[],
            si_units=units,
            electrode_attributes=schema,
            map_labels=normalize_map_labels(map_labels_payload, schema),
        )
    if spec == NATIVE_SPECIFICATION or isinstance(data.get("electrodes"), list):
        models, units, schema = _load_native_or_legacy_custom(data)
        return ArrayDocument(
            electrodes=models,
            pads=_load_pads_from_payload(data),
            orientation_markers=_load_orientation_markers_from_payload(data),
            si_units=units,
            electrode_attributes=schema,
            map_labels=normalize_map_labels(data.get("map_labels", _MISSING), schema),
        )

    raise ValueError("Unsupported file format.")


def load_array_from_file(path: str) -> tuple[list[Electrode], list[Pad], str, list[AttributeSpec]]:
    """
    Load electrodes, pads and the electrode-attribute schema from a JSON file.

    Supported formats:
    1) mea_editor native JSON (pads optional, absent on older files)
    2) legacy custom editor JSON
    3) probeinterface JSON (read without the probeinterface package)

    Returns:
        (electrodes, pads, si_units, electrode_attributes)

    Raises:
        ValueError: unsupported format or empty content.
        OSError / json.JSONDecodeError: file errors (propagated).
    """
    document = load_array_document(path)
    return document.electrodes, document.pads, document.si_units, document.electrode_attributes


def load_electrodes_from_file(path: str) -> tuple[list[Electrode], str]:
    """
    Load electrodes from a JSON file.

    Supported formats:
    1) mea_editor native JSON
    2) legacy custom editor JSON
    3) probeinterface JSON (read without the probeinterface package)

    Returns:
        (models, si_units)

    Raises:
        ValueError: unsupported format or empty content.
        OSError / json.JSONDecodeError: file errors (propagated).
    """
    models, _pads, units, _schema = load_array_from_file(path)
    return models, units


def is_probeinterface_file(path: str) -> bool:
    """True if the JSON file is a probeinterface document."""
    try:
        with open(path, "r", encoding="utf-8") as fh:
            data = json.load(fh)
    except (OSError, json.JSONDecodeError, UnicodeDecodeError):
        return False
    return isinstance(data, dict) and data.get("specification") == PROBEINTERFACE_SPEC


def _electrode_to_native_dict(model: Electrode, schema: list[AttributeSpec] | None = None) -> dict[str, Any]:
    payload = {
        "eid": int(model.eid),
        "x": float(model.x),
        "y": float(model.y),
        "radius": max(float(model.radius), MIN_RADIUS),
        "height": max(float(model.height), 0.0),
        "potentiostat_id": int(model.potentiostat_id),
        "intan_id": str(model.intan_id),
        "manufacturer_id": str(model.manufacturer_id),
        "contact_plane_axis": [float(v) for v in model.contact_plane_axis],
        "shank_id": str(model.shank_id),
        "shape": str(model.shape or DEFAULT_SHAPE),
        "label_position": normalize_label_position(model.label_position),
        "label_orientation": normalize_label_orientation(model.label_orientation),
    }
    extras = extra_specs(schema) if schema is not None else []
    attributes: dict[str, Any] = {}
    if extras:
        for spec in extras:
            raw = model.extra.get(spec.key, spec.default)
            attributes[spec.key] = coerce_value(spec.value_type, raw, spec.default)
    elif model.extra:
        attributes = dict(model.extra)
    if attributes:
        payload["attributes"] = attributes
    return payload


def _pad_to_native_dict(model: Pad) -> dict[str, Any]:
    return {
        "pad_id": int(model.pad_id),
        "electrode_eid": int(model.electrode_eid),
        "x": float(model.x),
        "y": float(model.y),
        "radius": max(float(model.radius), MIN_RADIUS),
        "height": max(float(model.height), 0.0),
        "shape": str(model.shape or DEFAULT_PAD_SHAPE),
        "label_position": normalize_label_position(model.label_position),
        "label_orientation": normalize_label_orientation(model.label_orientation),
    }


def _orientation_marker_to_native_dict(model: OrientationMarker) -> dict[str, Any]:
    return {
        "marker_id": int(model.marker_id),
        "x": float(model.x),
        "y": float(model.y),
        "radius": max(float(model.radius), MIN_RADIUS),
        "height": max(float(model.height), 0.0),
        "shape": str(model.shape or DEFAULT_MARKER_SHAPE),
        "label_position": normalize_label_position(model.label_position),
        "label_orientation": normalize_label_orientation(model.label_orientation),
    }


def save_electrodes_to_file(
    path: str,
    electrodes: list[Electrode],
    si_units: str,
    pads: list[Pad] | None = None,
    electrode_attributes: list[AttributeSpec] | None = None,
    map_labels: Iterable[str] | None = None,
    orientation_markers: list[OrientationMarker] | None = None,
) -> None:
    """
    Save electrodes (and optional pads / orientation markers) in native
    mea_editor JSON format.

    Extra electrode attributes and their schema are stored in the file so the
    editor can rebuild the same fields when the file is opened again.
    Pads, orientation markers, contact shapes, geometry (including rect
    height), visible map labels, and per-item label positions / orientations
    are persisted. Label position and orientation are native JSON only:
    SpikeInterface and XLSX exports omit them.
    """
    schema = list(electrode_attributes) if electrode_attributes is not None else schema_from_payload(None, electrodes)
    ordered = sorted(electrodes, key=lambda m: m.eid)
    ordered_pads = sorted(pads or [], key=lambda m: m.pad_id)
    ordered_markers = sorted(orientation_markers or [], key=lambda m: m.marker_id)
    payload = {
        "specification": NATIVE_SPECIFICATION,
        "version": NATIVE_VERSION,
        "editor_version": EDITOR_VERSION,
        "si_units": si_units or DEFAULT_UNITS,
        "electrode_attributes": schema_to_payload(schema),
        "map_labels": normalize_map_labels(
            None if map_labels is None else list(map_labels),
            schema,
        ),
        "electrodes": [_electrode_to_native_dict(m, schema) for m in ordered],
        "pads": [_pad_to_native_dict(m) for m in ordered_pads],
        "orientation_markers": [_orientation_marker_to_native_dict(m) for m in ordered_markers],
    }
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=4)
        fh.write("\n")


save_array_to_file = save_electrodes_to_file


def _contact_ids_for_spikeinterface(electrodes: list[Electrode]) -> list[str]:
    """
    Build unique Contact IDs for probeinterface.

    Prefer Manufacturer ID when every electrode has one; otherwise INTAN ID.
    """
    manufacturer_ids = [str(m.manufacturer_id).strip() for m in electrodes]
    if all(manufacturer_ids):
        return manufacturer_ids
    if any(manufacturer_ids):
        missing = [m.eid for m, mid in zip(electrodes, manufacturer_ids) if not mid]
        raise ValueError(
            "Manufacturer ID must be filled on every electrode (or left empty on all) "
            f"before SpikeInterface export. Missing on eid: {missing}."
        )
    intan_ids = [str(m.intan_id).strip() for m in electrodes]
    if any(not cid for cid in intan_ids):
        missing = [m.eid for m, cid in zip(electrodes, intan_ids) if not cid]
        raise ValueError(f"INTAN ID is empty on electrode eid: {missing}.")
    return intan_ids


def _pad_annotation_columns(
    ordered: list[Electrode],
    pad_by_eid: dict[int, Pad],
) -> dict[str, list[Any]]:
    """First linked pad per electrode, or empty cells when the electrode has none."""
    columns: dict[str, list[Any]] = {key: [] for key in PAD_GEOMETRY_ANNOTATION_KEYS}
    for model in ordered:
        pad = pad_by_eid.get(model.eid)
        if pad is None:
            for key in PAD_GEOMETRY_ANNOTATION_KEYS:
                columns[key].append("")
            continue
        columns["pad_id"].append(int(pad.pad_id))
        columns["pad_x"].append(float(pad.x))
        columns["pad_y"].append(float(pad.y))
        columns["pad_shape"].append(str(pad.shape or DEFAULT_PAD_SHAPE))
        columns["pad_radius"].append(max(float(pad.radius), MIN_RADIUS))
        columns["pad_height"].append(max(float(pad.height), 0.0))
    return columns


def build_probeinterface_payload(
    electrodes: list[Electrode],
    si_units: str,
    pads: list[Pad] | None = None,
    electrode_attributes: list[AttributeSpec] | None = None,
    map_labels: Iterable[str] | None = None,
) -> dict[str, Any]:
    """
    Build a probeinterface JSON dict for SpikeInterface.

    Channel ID  -> device_channel_indices (from INTAN ID)
    Contact ID  -> contact_ids (from Manufacturer ID, fallback INTAN ID)

    Native identifiers, extra attributes, and the first linked pad (id +
    geometry) are stored in contact_annotations. The file-level attribute
    schema and visible map labels are stored on the probe so a reopened
    export keeps extra-field meaning and map display. Pads themselves are
    not exported as SpikeInterface contacts. Per-item label position and
    orientation are omitted (native JSON only).
    """
    if not electrodes:
        raise ValueError("No electrodes to export.")

    ordered = sorted(electrodes, key=lambda m: m.eid)
    contact_ids = _contact_ids_for_spikeinterface(ordered)
    unique_contacts = set(contact_ids)
    if len(unique_contacts) != len(contact_ids):
        raise ValueError(
            "Contact IDs must be unique for SpikeInterface / probeinterface. "
            "Fill unique Manufacturer IDs (or unique INTAN IDs if Manufacturer ID is unused)."
        )

    channel_ids: list[int] = []
    errors: list[str] = []
    for model in ordered:
        try:
            channel_ids.append(intan_id_to_channel_id(model.intan_id))
        except ValueError as exc:
            errors.append(f"eid {model.eid}: {exc}")
    if errors:
        raise ValueError("Could not derive channel IDs from INTAN IDs:\n" + "\n".join(errors))

    duplicate_channels = sorted(
        channel for channel, count in Counter(ch for ch in channel_ids if ch >= 0).items() if count > 1
    )
    if duplicate_channels:
        raise ValueError(
            "Channel IDs derived from INTAN ID must be unique (except NC / disconnected = -1). "
            f"Duplicates: {duplicate_channels}."
        )

    schema = (
        list(electrode_attributes)
        if electrode_attributes is not None
        else schema_from_payload(None, ordered)
    )
    fill_electrodes_extras(ordered, schema, prune=False)
    pad_by_eid = _first_pad_by_electrode(pads)
    contact_annotations: dict[str, list[Any]] = {
        SI_CHANNEL_ID_KEY: channel_ids,
        SI_CONTACT_ID_KEY: contact_ids,
        "eid": [int(model.eid) for model in ordered],
        "potentiostat_id": [int(model.potentiostat_id) for model in ordered],
        "intan_id": [str(model.intan_id) for model in ordered],
        "manufacturer_id": [str(model.manufacturer_id) for model in ordered],
        "shank_id": [str(model.shank_id) for model in ordered],
    }
    contact_annotations.update(_pad_annotation_columns(ordered, pad_by_eid))
    extras = extra_specs(schema)
    for spec in extras:
        contact_annotations[spec.key] = [
            coerce_value(spec.value_type, model.get_attribute(spec.key), spec.default)
            for model in ordered
        ]

    probe: dict[str, Any] = {
        "ndim": 2,
        "si_units": si_units_for_probeinterface(si_units),
        "annotations": {
            "name": "mea_editor",
            "model_name": "mea_editor",
            "manufacturer": "mea_editor",
            MEA_EDITOR_SCHEMA_ANNOTATION_KEY: schema_to_payload(schema),
            MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY: normalize_map_labels(
                None if map_labels is None else list(map_labels),
                schema,
            ),
        },
        "contact_annotations": contact_annotations,
        "contact_positions": [[float(m.x), float(m.y)] for m in ordered],
        "contact_plane_axes": [
            [
                [float(m.contact_plane_axis[0]), float(m.contact_plane_axis[1])],
                [float(m.contact_plane_axis[2]), float(m.contact_plane_axis[3])],
            ]
            for m in ordered
        ],
        "contact_shapes": [
            normalize_contact_shape(str(m.shape or DEFAULT_SHAPE), DEFAULT_ELECTRODE_SHAPE)
            for m in ordered
        ],
        "contact_shape_params": [
            probe_shape_params(m.shape, m.radius, m.height, MIN_RADIUS) for m in ordered
        ],
        "device_channel_indices": channel_ids,
        "contact_ids": contact_ids,
        "shank_ids": [str(m.shank_id) for m in ordered],
    }
    return {
        "specification": PROBEINTERFACE_SPEC,
        "version": PROBEINTERFACE_VERSION,
        "probes": [probe],
    }


def export_spikeinterface_json(
    path: str,
    electrodes: list[Electrode],
    si_units: str,
    pads: list[Pad] | None = None,
    electrode_attributes: list[AttributeSpec] | None = None,
    map_labels: Iterable[str] | None = None,
) -> None:
    """
    Write a probeinterface JSON file usable by SpikeInterface.

    The JSON contains:
    - device_channel_indices: channel ID (from INTAN ID)
    - contact_ids: Contact ID (from Manufacturer ID, fallback INTAN ID)
    - contact_annotations: native IDs (including shank_id), extra attributes,
      linked pad id and geometry
    - probe annotations: electrode attribute schema and visible map labels
    """
    payload = build_probeinterface_payload(
        electrodes,
        si_units,
        pads=pads,
        electrode_attributes=electrode_attributes,
        map_labels=map_labels,
    )
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=4)
        fh.write("\n")


def _require_openpyxl_workbook():
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise ImportError("openpyxl is required for XLSX export.") from exc
    return Workbook


def _resolved_schema(
    electrodes: list[Electrode],
    electrode_attributes: list[AttributeSpec] | None,
) -> list[AttributeSpec]:
    if electrode_attributes is not None:
        schema = list(electrode_attributes)
    else:
        schema = schema_from_payload(None, electrodes)
    fill_electrodes_extras(electrodes, schema, prune=False)
    return schema


def _write_orientation_marker_sheet(workbook, orientation_markers: list[OrientationMarker] | None) -> None:
    """Add an orientation_markers sheet (id, x, y, shape, sizes). Always created."""
    sheet = workbook.create_sheet("orientation_markers")
    sheet.append(["marker_id", "x", "y", "shape", "radius", "width", "height"])
    for marker in sorted(orientation_markers or [], key=lambda item: item.marker_id):
        radius, width, height = export_contact_sizes(marker.shape, marker.radius, marker.height)
        sheet.append(
            [
                int(marker.marker_id),
                float(marker.x),
                float(marker.y),
                str(marker.shape or DEFAULT_MARKER_SHAPE),
                radius,
                width,
                height,
            ]
        )


def export_analysis_xlsx(
    path: str,
    electrodes: list[Electrode],
    pads: list[Pad] | None = None,
    electrode_attributes: list[AttributeSpec] | None = None,
    orientation_markers: list[OrientationMarker] | None = None,
) -> None:
    """
    Write the analysis table used by downstream mapping scripts.

    The first four columns stay stable:
    - channel: Potentiostat ID
    - row: electrode y
    - col: electrode x
    - shape: electrode shape

    Then radius / width / height (circle: radius; square: width and height
    equal to the side; rect: independent width and height), INTAN ID,
    `si_channel` (SpikeInterface channel derived from INTAN; empty if invalid),
    manufacturer / shank / eid, extra attributes, and the first linked pad
    (`pad_id`, `pad_x`, `pad_y`, `pad_shape`).

    Orientation markers are written on a separate `orientation_markers` sheet
    (`marker_id`, `x`, `y`, `shape`, `radius`, `width`, `height`). They are
    not SpikeInterface contacts.

    `channel` is the Potentiostat ID, not the SpikeInterface channel.
    """
    if not electrodes:
        raise ValueError("No electrodes to export.")
    Workbook = _require_openpyxl_workbook()
    schema = _resolved_schema(electrodes, electrode_attributes)
    extras = extra_specs(schema)
    pad_by_eid = _first_pad_by_electrode(pads)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "array"
    headers = [
        "channel",
        "row",
        "col",
        "shape",
        "radius",
        "width",
        "height",
        "intan_id",
        "si_channel",
        "manufacturer_id",
        "shank_id",
        "eid",
        *[spec.key for spec in extras],
        "pad_id",
        "pad_x",
        "pad_y",
        "pad_shape",
    ]
    worksheet.append(headers)
    for model in sorted(electrodes, key=lambda item: (item.potentiostat_id, item.eid)):
        pad = pad_by_eid.get(model.eid)
        radius, width, height = export_contact_sizes(model.shape, model.radius, model.height)
        si_channel = try_intan_channel_id(model.intan_id)
        row = [
            model.potentiostat_id,
            model.y,
            model.x,
            model.shape,
            radius,
            width,
            height,
            model.intan_id,
            "" if si_channel is None else si_channel,
            model.manufacturer_id,
            model.shank_id,
            int(model.eid),
        ]
        row.extend(_extra_values(model, extras))
        if pad is None:
            row.extend(["", "", "", ""])
        else:
            row.extend([int(pad.pad_id), float(pad.x), float(pad.y), pad.shape])
        worksheet.append(row)
    _write_orientation_marker_sheet(workbook, orientation_markers)
    workbook.save(path)
    workbook.close()


def export_array_xlsx(
    path: str,
    electrodes: list[Electrode],
    pads: list[Pad] | None = None,
    electrode_attributes: list[AttributeSpec] | None = None,
    si_units: str = DEFAULT_UNITS,
    orientation_markers: list[OrientationMarker] | None = None,
) -> None:
    """
    Write a full array workbook: electrodes, pads, orientation markers, schema.

    Extra electrode attributes follow the file schema on both the electrode
    sheet and the linked-electrode columns of the pads sheet. Orientation
    markers are not linked to electrodes or pads.
    """
    if not electrodes:
        raise ValueError("No electrodes to export.")
    Workbook = _require_openpyxl_workbook()
    schema = _resolved_schema(electrodes, electrode_attributes)
    extras = extra_specs(schema)
    pad_by_eid = _first_pad_by_electrode(pads)
    electrode_by_eid = {model.eid: model for model in electrodes}

    workbook = Workbook()
    electrodes_sheet = workbook.active
    electrodes_sheet.title = "array"
    electrode_headers = [
        "eid",
        "potentiostat_id",
        "intan_id",
        "si_channel",
        "manufacturer_id",
        "row",
        "col",
        "shank_id",
        "shape",
        "radius",
        "width",
        "height",
        "contact_plane_axis",
        *[spec.key for spec in extras],
        "pad_id",
        "si_units",
    ]
    electrodes_sheet.append(electrode_headers)
    for model in sorted(electrodes, key=lambda item: (item.potentiostat_id, item.eid)):
        pad = pad_by_eid.get(model.eid)
        radius, width, height = export_contact_sizes(model.shape, model.radius, model.height)
        si_channel = try_intan_channel_id(model.intan_id)
        row = [
            int(model.eid),
            model.potentiostat_id,
            model.intan_id,
            "" if si_channel is None else si_channel,
            model.manufacturer_id,
            model.y,
            model.x,
            model.shank_id,
            model.shape,
            radius,
            width,
            height,
            _contact_plane_axis_text(model.contact_plane_axis),
        ]
        row.extend(_extra_values(model, extras))
        row.extend(
            [
                int(pad.pad_id) if pad is not None else "",
                si_units or DEFAULT_UNITS,
            ]
        )
        electrodes_sheet.append(row)

    pads_sheet = workbook.create_sheet("pads")
    pads_sheet.append(
        [
            "pad_id",
            "electrode_eid",
            "potentiostat_id",
            "intan_id",
            "si_channel",
            "manufacturer_id",
            "shank_id",
            "x",
            "y",
            "shape",
            "radius",
            "width",
            "height",
            *[spec.key for spec in extras],
        ]
    )
    for pad in sorted(pads or [], key=lambda item: item.pad_id):
        electrode = electrode_by_eid.get(pad.electrode_eid)
        radius, width, height = export_contact_sizes(pad.shape, pad.radius, pad.height)
        si_channel = (
            try_intan_channel_id(electrode.intan_id) if electrode is not None else None
        )
        row = [
            int(pad.pad_id),
            int(pad.electrode_eid),
            electrode.potentiostat_id if electrode is not None else "",
            electrode.intan_id if electrode is not None else "",
            "" if si_channel is None else si_channel,
            electrode.manufacturer_id if electrode is not None else "",
            electrode.shank_id if electrode is not None else "",
            float(pad.x),
            float(pad.y),
            pad.shape,
            radius,
            width,
            height,
        ]
        row.extend(_extra_values(electrode, extras))
        pads_sheet.append(row)

    _write_orientation_marker_sheet(workbook, orientation_markers)

    schema_sheet = workbook.create_sheet("electrode_attributes")
    schema_sheet.append(["key", "label", "type", "default", "builtin", "unique", "unique_scope"])
    for spec in schema:
        schema_sheet.append(
            [
                spec.key,
                spec.label,
                spec.value_type,
                spec.default,
                bool(spec.builtin),
                bool(spec.unique),
                spec.unique_scope,
            ]
        )
    workbook.save(path)
    workbook.close()
