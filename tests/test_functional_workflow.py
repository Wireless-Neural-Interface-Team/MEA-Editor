"""End-to-end check of editor workflows: new array, edit, add, I/O, table, undo."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtWidgets import QApplication, QMessageBox

from mea_editor.array_integrity import pairing_problems
from mea_editor.attribute_schema import AttributeSpec
from mea_editor.contact_shape import size_field_from_stored_half
from mea_editor.electrode import DEFAULT_MAP_LABEL_KEYS
from mea_editor.electrode_array_dialogs import AddAttributeDialog, NewArrayDialog, NewArrayParams
from mea_editor.electrode_array_editor_io import NATIVE_SPECIFICATION, NATIVE_VERSION, load_array_document
from mea_editor.electrode_array_editor_qt import ElectrodeArrayEditorQt
from mea_editor.electrode_table_window import EID_ROLE


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SAMPLE_PROBEINTERFACE = PROJECT_ROOT / "MEA_RdLGN64test.json"


def _pump(times: int = 8) -> None:
    app = QApplication.instance()
    assert app is not None
    for _ in range(times):
        app.processEvents()


class FunctionalWorkflowTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls._app = QApplication.instance() or QApplication([])

    def setUp(self) -> None:
        self.editor = ElectrodeArrayEditorQt()
        self.editor._startup_done = True
        self._orig_info = QMessageBox.information
        self._orig_warn = QMessageBox.warning
        self._orig_crit = QMessageBox.critical
        self._orig_question = QMessageBox.question
        QMessageBox.information = staticmethod(lambda *args, **kwargs: QMessageBox.Ok)
        QMessageBox.warning = staticmethod(lambda *args, **kwargs: QMessageBox.Ok)
        QMessageBox.critical = staticmethod(lambda *args, **kwargs: QMessageBox.Ok)
        QMessageBox.question = staticmethod(lambda *args, **kwargs: QMessageBox.Yes)

    def tearDown(self) -> None:
        QMessageBox.information = self._orig_info
        QMessageBox.warning = self._orig_warn
        QMessageBox.critical = self._orig_crit
        QMessageBox.question = self._orig_question
        if self.editor._electrode_table_window is not None:
            self.editor._electrode_table_window.hide()
            self.editor._electrode_table_window.close()
        self.editor.is_dirty = False
        self.editor.close()
        _pump()

    def _make_grid(self, rows: int = 2, cols: int = 2) -> None:
        params = NewArrayParams(
            rows=rows,
            cols=cols,
            pitch=50.0,
            si_units="um",
            electrode_shape="circle",
            pad_size=10.0,
            pad_height=0.0,
            pad_shape="square",
            pad_rows=1,
            pad_spacing=50.0,
        )
        self.editor._generate_aligned_grid(params)
        _pump()

    def _select_electrode(self, eid: int) -> None:
        self.editor.scene.clearSelection()
        self.editor.items[eid].setSelected(True)
        _pump()
        self.editor._refresh_panel_values()

    def test_new_array_creates_paired_grid_and_fits_views(self) -> None:
        self._make_grid(2, 3)
        self.assertEqual(len(self.editor.electrodes), 6)
        self.assertEqual(len(self.editor.pads), 6)
        self.assertEqual(pairing_problems(self.editor.electrodes.values(), self.editor.pads.values()), [])
        xs = sorted({round(m.x, 6) for m in self.editor.electrodes.values()})
        ys = sorted({round(m.y, 6) for m in self.editor.electrodes.values()})
        self.assertEqual(xs, [0.0, 50.0, 100.0])
        self.assertEqual(ys, [0.0, 50.0])
        self.assertEqual(self.editor.si_units, "um")
        self.assertEqual(self.editor.visible_map_label_keys, set(DEFAULT_MAP_LABEL_KEYS))
        for model in self.editor.electrodes.values():
            self.assertEqual(model.shape, "circle")
            self.assertEqual(model.label_position, "below")
        for pad in self.editor.pads.values():
            self.assertEqual(pad.shape, "square")
            self.assertAlmostEqual(pad.radius, 10.0)
        self.editor._fit_all_views()
        _pump()
        self.assertTrue(self.editor.electrode_view.sceneRect().isValid())
        self.assertTrue(self.editor.pads_map_view.sceneRect().isValid())

    def test_new_array_dialog_values_match_widgets(self) -> None:
        dialog = NewArrayDialog()
        dialog.rows_spin.setValue(4)
        dialog.cols_spin.setValue(5)
        dialog.pitch_spin.setValue(80.0)
        dialog.units_edit.setText("mm")
        dialog.electrode_shape_combo.setCurrentText("square")
        dialog.pad_shape_combo.setCurrentText("rect")
        dialog.pad_size_spin.setValue(12.0)
        dialog.pad_height_spin.setValue(6.0)
        dialog.pad_rows_spin.setValue(2)
        dialog.pad_spacing_spin.setValue(40.0)
        params = dialog.values()
        self.assertEqual(params.rows, 4)
        self.assertEqual(params.cols, 5)
        self.assertAlmostEqual(params.pitch, 80.0)
        self.assertEqual(params.si_units, "mm")
        self.assertEqual(params.electrode_shape, "square")
        self.assertEqual(params.pad_shape, "rect")
        self.assertAlmostEqual(params.pad_size, 6.0)
        self.assertAlmostEqual(params.pad_height, 3.0)
        self.assertEqual(params.pad_rows, 2)
        self.assertAlmostEqual(params.pad_spacing, 40.0)
        dialog.close()

    def test_selection_syncs_electrode_and_pad(self) -> None:
        self._make_grid()
        eid = 0
        self._select_electrode(eid)
        self.assertTrue(self.editor.items[eid].isSelected())
        linked = [p for p in self.editor.pads.values() if p.electrode_eid == eid]
        self.assertEqual(len(linked), 1)
        self.assertTrue(self.editor.pad_items[linked[0].pad_id].isSelected())
        self.assertIn("electrode", self.editor.selected_count_label.text().lower())

    def test_apply_electrode_and_pad_edits(self) -> None:
        self._make_grid()
        self._select_electrode(0)
        self.editor.shape_combo.setCurrentText("rect")
        self.editor.radius_edit.setText("24")
        self.editor.height_edit.setText("10")
        self.editor.x_edit.setText("7.5")
        self.editor.y_edit.setText("8.5")
        self.editor.attribute_edits["potentiostat_id"].setText("42")
        self.editor.attribute_edits["intan_id"].setText("B-001")
        self.editor.attribute_edits["manufacturer_id"].setText("M42")
        self.editor.attribute_edits["shank_id"].setText("3")
        self.editor._set_label_position_combo(self.editor.label_position_combo, "left", mixed=False)
        self.editor._set_label_orientation_combo(self.editor.label_orientation_combo, 90, mixed=False)
        self.editor._apply_pending_edits()
        _pump()
        model = self.editor.electrodes[0]
        self.assertEqual(model.shape, "rect")
        self.assertAlmostEqual(model.radius, 12.0)
        self.assertAlmostEqual(model.height, 5.0)
        self.assertAlmostEqual(model.x, 7.5)
        self.assertAlmostEqual(model.y, 8.5)
        self.assertEqual(model.potentiostat_id, 42)
        self.assertEqual(model.intan_id, "B-001")
        self.assertEqual(model.manufacturer_id, "M42")
        self.assertEqual(model.shank_id, "3")
        self.assertEqual(model.label_position, "left")
        self.assertEqual(model.label_orientation, 90)

        pad = next(p for p in self.editor.pads.values() if p.electrode_eid == 0)
        self.editor.scene.clearSelection()
        self.editor.pad_items[pad.pad_id].setSelected(True)
        _pump()
        self.editor._refresh_panel_values()
        self.editor.pad_shape_combo.setCurrentText("circle")
        self.editor.pad_radius_edit.setText("15")
        self.editor.pad_x_edit.setText("-40")
        self.editor.pad_y_edit.setText("12")
        self.editor._set_label_position_combo(self.editor.pad_label_position_combo, "above", mixed=False)
        self.editor._set_label_orientation_combo(self.editor.pad_label_orientation_combo, 180, mixed=False)
        self.editor._apply_pending_pad_edits()
        _pump()
        pad = self.editor.pads[pad.pad_id]
        self.assertEqual(pad.shape, "circle")
        self.assertAlmostEqual(pad.radius, 15.0)
        self.assertAlmostEqual(pad.x, -40.0)
        self.assertAlmostEqual(pad.y, 12.0)
        self.assertEqual(pad.label_position, "above")
        self.assertEqual(pad.label_orientation, 180)

    def test_dx_dy_moves_user_selection_only(self) -> None:
        self._make_grid()
        self._select_electrode(0)
        electrode_before = (self.editor.electrodes[0].x, self.editor.electrodes[0].y)
        pad = next(p for p in self.editor.pads.values() if p.electrode_eid == 0)
        pad_before = (pad.x, pad.y)
        self.editor.dx_edit.setText("5")
        self.editor.dy_edit.setText("-3")
        self.editor._move_selection_by_delta()
        _pump()
        self.assertAlmostEqual(self.editor.electrodes[0].x, electrode_before[0] + 5.0)
        self.assertAlmostEqual(self.editor.electrodes[0].y, electrode_before[1] - 3.0)
        pad = self.editor.pads[pad.pad_id]
        self.assertAlmostEqual(pad.x, pad_before[0])
        self.assertAlmostEqual(pad.y, pad_before[1])

    def test_add_electrode_pad_and_marker_then_undo_redo(self) -> None:
        self._make_grid()
        n_e = len(self.editor.electrodes)
        n_p = len(self.editor.pads)
        self.editor.add_electrode_shape_combo.setCurrentText("square")
        self.editor.add_electrode_size_edit.setText("20")
        self.editor._add_electrode_at(200.0, 200.0)
        _pump()
        self.assertEqual(len(self.editor.electrodes), n_e + 1)
        self.assertEqual(len(self.editor.pads), n_p)
        new_eid = max(self.editor.electrodes)
        self.assertEqual(self.editor.electrodes[new_eid].shape, "square")
        self.assertAlmostEqual(self.editor.electrodes[new_eid].radius, 10.0)

        self.editor.pad_add_electrode_combo.setCurrentIndex(
            self.editor.pad_add_electrode_combo.findData(new_eid)
        )
        self.editor.add_pad_shape_combo.setCurrentText("rect")
        self.editor.add_pad_size_edit.setText("16")
        self.editor.add_pad_height_edit.setText("8")
        self.editor._add_pad_at(250.0, 210.0)
        _pump()
        self.assertEqual(len(self.editor.pads), n_p + 1)
        new_pad = next(p for p in self.editor.pads.values() if p.electrode_eid == new_eid)
        self.assertEqual(new_pad.shape, "rect")
        self.assertAlmostEqual(new_pad.radius, 8.0)
        self.assertAlmostEqual(new_pad.height, 4.0)

        self.editor.add_marker_shape_combo.setCurrentText("circle")
        self.editor.add_marker_size_edit.setText("30")
        self.editor._add_marker_at(-20.0, -20.0)
        _pump()
        self.assertEqual(len(self.editor.orientation_markers), 1)
        marker = next(iter(self.editor.orientation_markers.values()))
        self.assertEqual(marker.shape, "circle")
        self.assertAlmostEqual(marker.radius, 30.0)

        self.editor.scene.clearSelection()
        self.editor.marker_items[marker.marker_id].setSelected(True)
        _pump()
        self.editor._refresh_panel_values()
        self.editor.marker_shape_combo.setCurrentText("square")
        self.editor.marker_radius_edit.setText("40")
        self.editor.marker_x_edit.setText("-25")
        self.editor.marker_y_edit.setText("-30")
        self.editor._apply_pending_marker_edits()
        _pump()
        marker = self.editor.orientation_markers[marker.marker_id]
        self.assertEqual(marker.shape, "square")
        self.assertAlmostEqual(marker.radius, 20.0)
        self.assertAlmostEqual(marker.x, -25.0)
        self.assertAlmostEqual(marker.y, -30.0)

        self.editor._undo()
        _pump()
        marker = next(iter(self.editor.orientation_markers.values()))
        self.assertEqual(marker.shape, "circle")
        self.editor._redo()
        _pump()
        marker = next(iter(self.editor.orientation_markers.values()))
        self.assertEqual(marker.shape, "square")

    def test_map_labels_and_extra_attributes(self) -> None:
        self._make_grid()
        self.assertTrue("intan_id" in self.editor.visible_map_label_keys)
        self.editor._set_map_label_visible("intan_id", False)
        self.assertNotIn("intan_id", self.editor.visible_map_label_keys)
        self.assertFalse(self.editor._map_label_checks["intan_id"].isChecked())
        self.editor._set_map_label_visible("manufacturer_id", True)
        self.assertIn("manufacturer_id", self.editor.visible_map_label_keys)

        dialog = AddAttributeDialog(self.editor.attribute_schema)
        dialog.name_edit.setText("Site note")
        dialog.type_combo.setCurrentText("str")
        spec = dialog.spec()
        dialog.close()
        self.assertIsNotNone(spec)
        assert spec is not None
        schema = list(self.editor.attribute_schema) + [spec]
        self.editor._set_attribute_schema(schema, prune=False)
        _pump()
        self.assertIn("site_note", self.editor.attribute_edits)
        self._select_electrode(0)
        self.editor.attribute_edits["site_note"].setText("deep")
        self.editor._apply_pending_edits()
        self.assertEqual(self.editor.electrodes[0].extra.get("site_note"), "deep")
        self.editor._set_map_label_visible("site_note", True)
        inner, outer = self.editor._labels_for_electrode(self.editor.electrodes[0])
        self.assertIn("deep", " ".join((inner, outer)))

    def test_delete_electrode_removes_linked_pad(self) -> None:
        self._make_grid()
        self._select_electrode(0)
        self.editor._delete_selected()
        _pump()
        self.assertNotIn(0, self.editor.electrodes)
        self.assertFalse(any(p.electrode_eid == 0 for p in self.editor.pads.values()))
        self.assertEqual(len(self.editor.electrodes), 3)
        self.assertEqual(len(self.editor.pads), 3)

    def test_delete_pad_only_keeps_electrode(self) -> None:
        self._make_grid()
        pad = next(iter(self.editor.pads.values()))
        eid = pad.electrode_eid
        self.editor.scene.clearSelection()
        self.editor.pad_items[pad.pad_id].setSelected(True)
        _pump()
        self.editor._refresh_panel_values()
        self.editor._delete_selected()
        _pump()
        self.assertIn(eid, self.editor.electrodes)
        self.assertNotIn(pad.pad_id, self.editor.pads)
        problems = pairing_problems(self.editor.electrodes.values(), self.editor.pads.values())
        self.assertTrue(problems)

    def test_find_pad_and_marker(self) -> None:
        self._make_grid()
        self.editor._add_marker_at(0.0, -80.0)
        _pump()
        pad_id = min(self.editor.pads)
        self.editor.pad_id_find_edit.setText(str(pad_id))
        self.editor._find_by_pad_id()
        _pump()
        self.assertTrue(self.editor.pad_items[pad_id].isSelected())
        marker_id = min(self.editor.orientation_markers)
        self.editor.marker_id_find_edit.setText(str(marker_id))
        self.editor._find_by_marker_id()
        _pump()
        self.assertTrue(self.editor.marker_items[marker_id].isSelected())

    def test_find_by_attributes(self) -> None:
        self._make_grid()
        self.editor.attribute_find_edit.setText("A-000")
        self.editor._find_by_attributes()
        _pump()
        matches = [m for m in self.editor.electrodes.values() if m.intan_id == "A-000"]
        self.assertEqual(len(matches), 1)
        self.assertTrue(self.editor.items[matches[0].eid].isSelected())

    def test_duplicate_intan_is_flagged(self) -> None:
        self._make_grid()
        self._select_electrode(1)
        self.editor.attribute_edits["intan_id"].setText("A-000")
        self.editor._apply_pending_edits()
        _pump()
        self.assertTrue(self.editor.electrodes[0].has_intan_duplicate)
        self.assertTrue(self.editor.electrodes[1].has_intan_duplicate)

    def test_electrode_table_search_and_row_selection(self) -> None:
        self._make_grid()
        self.editor._show_electrode_table()
        _pump()
        win = self.editor._electrode_table_window
        self.assertIsNotNone(win)
        assert win is not None
        self.assertTrue(win.isVisible())
        self.assertEqual(win.proxy.rowCount(), 4)
        win.search_edit.setText("A-002")
        _pump()
        self.assertEqual(win.proxy.rowCount(), 1)
        eid = int(win.proxy.index(0, 0).data(EID_ROLE))
        win.set_selected_eids([eid])
        self.editor._select_from_electrode_table([eid])
        _pump()
        self.assertTrue(self.editor.items[eid].isSelected())
        win.search_edit.setText("")
        _pump()
        self.assertEqual(win.proxy.rowCount(), 4)
        win.hide()

    def test_native_save_load_and_exports(self) -> None:
        self._make_grid()
        self.editor._add_marker_at(-10.0, -10.0)
        schema = list(self.editor.attribute_schema) + [
            AttributeSpec(key="site_note", label="Site note", value_type="str", default="")
        ]
        self.editor._set_attribute_schema(schema, prune=False)
        self.editor.electrodes[0].extra["site_note"] = "deep"
        self.editor._set_map_label_visible("site_note", True)
        self.editor.si_units_edit.setText("mm")
        self.editor._apply_si_units()

        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            folder = Path(tmp)
            native = folder / "array.json"
            si_path = folder / "si.json"
            analysis = folder / "analysis.xlsx"
            full = folder / "array.xlsx"
            self.editor._save_array_to_file(str(native))
            payload = json.loads(native.read_text(encoding="utf-8"))
            self.assertEqual(payload["specification"], NATIVE_SPECIFICATION)
            self.assertEqual(payload["version"], NATIVE_VERSION)
            self.assertEqual(payload["si_units"], "mm")
            self.assertEqual(len(payload["electrodes"]), 4)
            self.assertEqual(len(payload["pads"]), 4)
            self.assertEqual(len(payload["orientation_markers"]), 1)
            extra_keys = {item["key"] for item in payload["electrode_attributes"]}
            self.assertIn("site_note", extra_keys)

            self.editor._set_array([], [])
            self.editor._load_array_from_file(str(native))
            _pump()
            self.assertEqual(len(self.editor.electrodes), 4)
            self.assertEqual(len(self.editor.pads), 4)
            self.assertEqual(len(self.editor.orientation_markers), 1)
            self.assertEqual(self.editor.si_units, "mm")
            self.assertEqual(self.editor.electrodes[0].extra.get("site_note"), "deep")
            self.assertIn("site_note", self.editor.visible_map_label_keys)

            self.editor._export_analysis_to_xlsx(str(analysis))
            self.editor._export_matrix_to_xlsx(str(full))
            from mea_editor.electrode_array_editor_io import export_spikeinterface_json

            export_spikeinterface_json(
                str(si_path),
                list(self.editor.electrodes.values()),
                self.editor.si_units,
                pads=list(self.editor.pads.values()),
                electrode_attributes=self.editor.attribute_schema,
                map_labels=self.editor.visible_map_label_keys,
            )
            si_payload = json.loads(si_path.read_text(encoding="utf-8"))
            self.assertEqual(si_payload["specification"], "probeinterface")
            self.assertEqual(len(si_payload["probes"][0]["contact_positions"]), 4)

            analysis_wb = load_workbook(analysis)
            try:
                self.assertIn("array", analysis_wb.sheetnames)
                self.assertIn("orientation_markers", analysis_wb.sheetnames)
                headers = [cell.value for cell in analysis_wb["array"][1]]
                self.assertIn("channel", headers)
                self.assertIn("si_channel", headers)
                self.assertEqual(analysis_wb["orientation_markers"].max_row, 2)
            finally:
                analysis_wb.close()

            full_wb = load_workbook(full)
            try:
                for name in ("array", "pads", "orientation_markers", "electrode_attributes"):
                    self.assertIn(name, full_wb.sheetnames)
                self.assertEqual(full_wb["array"].max_row, 5)
                self.assertEqual(full_wb["pads"].max_row, 5)
            finally:
                full_wb.close()

    def test_open_probeinterface_sample(self) -> None:
        self.assertTrue(SAMPLE_PROBEINTERFACE.is_file())
        self.editor._load_array_from_file(str(SAMPLE_PROBEINTERFACE))
        _pump()
        self.assertGreaterEqual(len(self.editor.electrodes), 1)
        document = load_array_document(str(SAMPLE_PROBEINTERFACE))
        self.assertEqual(len(self.editor.electrodes), len(document.electrodes))
        self.editor._fit_all_views()
        _pump()
        self.assertTrue(self.editor.electrode_view.transform().isInvertible())

    def test_si_units_and_menus_exist(self) -> None:
        self._make_grid()
        self.editor.si_units_edit.setText("mm")
        self.editor._apply_si_units()
        self.assertEqual(self.editor.si_units, "mm")
        menu_actions = list(self.editor.menuBar().actions())
        titles = [action.text() for action in menu_actions]
        self.assertEqual(titles, ["File", "Edit", "View", "Help"])
        file_menu = menu_actions[0].menu()
        self.assertIsNotNone(file_menu)
        assert file_menu is not None
        file_texts = [action.text() for action in file_menu.actions() if action.text()]
        for needed in (
            "New array...",
            "Open...",
            "Save",
            "Save As...",
            "Export for SpikeInterface...",
            "Export for analysis...",
            "Export array as XLSX...",
        ):
            self.assertIn(needed, file_texts)

    def test_add_mode_escape_and_size_labels(self) -> None:
        self._make_grid()
        self.editor.b_add_electrode.setChecked(True)
        self.assertTrue(self.editor.is_add_mode)
        self.editor._stop_all_add_modes()
        self.assertFalse(self.editor.is_add_mode)
        self.editor.shape_combo.setCurrentText("square")
        self.assertEqual(self.editor.electrode_size_label.text(), "Side length")
        first = next(iter(self.editor.electrodes.values()))
        shown = size_field_from_stored_half(first.shape, first.radius)
        self.assertGreater(shown, 0.0)


if __name__ == "__main__":
    unittest.main()
