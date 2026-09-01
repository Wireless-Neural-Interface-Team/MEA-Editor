"""Round-trip checks for native save, SpikeInterface export, and XLSX."""

from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from mea_editor import __version__
from mea_editor.attribute_schema import AttributeSpec, default_schema
from mea_editor.contact_shape import export_contact_sizes
from mea_editor.electrode import DEFAULT_MAP_LABEL_KEYS, Electrode
from mea_editor.electrode_array_editor_io import (
    MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY,
    MEA_EDITOR_SCHEMA_ANNOTATION_KEY,
    NATIVE_SPECIFICATION,
    NATIVE_VERSION,
    build_probeinterface_payload,
    export_analysis_xlsx,
    export_array_xlsx,
    export_spikeinterface_json,
    load_array_document,
    load_array_from_file,
    save_array_to_file,
)
from mea_editor.orientation_marker import OrientationMarker
from mea_editor.pad import Pad


def _sample_array() -> tuple[list[Electrode], list[Pad], list[AttributeSpec]]:
    schema = default_schema() + [
        AttributeSpec(
            key="site_note",
            label="Site note",
            value_type="str",
            default="",
            unique=True,
            unique_scope="per_shank",
        )
    ]
    electrodes = [
        Electrode(
            eid=2,
            x=10.0,
            y=20.0,
            radius=12.0,
            height=8.0,
            potentiostat_id=7,
            intan_id="A-003",
            manufacturer_id="M3",
            shank_id="1",
            shape="rect",
            extra={"site_note": "deep"},
        ),
        Electrode(
            eid=5,
            x=40.0,
            y=20.0,
            radius=6.0,
            potentiostat_id=8,
            intan_id="A-004",
            manufacturer_id="M4",
            shape="square",
            extra={"site_note": "shallow"},
        ),
    ]
    pads = [
        Pad(
            pad_id=1,
            electrode_eid=2,
            x=-30.0,
            y=20.0,
            radius=10.0,
            height=4.0,
            shape="rect",
        ),
        Pad(
            pad_id=4,
            electrode_eid=5,
            x=80.0,
            y=20.0,
            radius=10.0,
            shape="square",
        ),
    ]
    return electrodes, pads, schema


class IoRoundTripTests(unittest.TestCase):
    def test_native_save_roundtrip(self) -> None:
        electrodes, pads, schema = _sample_array()
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "array.json")
            save_array_to_file(
                path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
                map_labels=["intan_id", "site_note"],
            )
            with open(path, encoding="utf-8") as fh:
                payload = json.load(fh)
            self.assertEqual(payload["specification"], NATIVE_SPECIFICATION)
            self.assertEqual(payload["version"], NATIVE_VERSION)
            self.assertEqual(payload["editor_version"], __version__)
            self.assertEqual(payload["map_labels"], ["intan_id", "site_note"])
            self.assertEqual(len(payload["electrodes"]), 2)
            self.assertEqual(len(payload["pads"]), 2)
            document = load_array_document(path)
        loaded, loaded_pads, units, loaded_schema = (
            document.electrodes,
            document.pads,
            document.si_units,
            document.electrode_attributes,
        )
        self.assertEqual(document.map_labels, ["intan_id", "site_note"])
        self.assertEqual(units, "um")
        self.assertEqual([m.eid for m in loaded], [2, 5])
        self.assertEqual(loaded[0].shape, "rect")
        self.assertEqual(loaded[0].height, 8.0)
        self.assertEqual(loaded[0].intan_id, "A-003")
        self.assertEqual(loaded[0].manufacturer_id, "M3")
        self.assertEqual(loaded[0].extra.get("site_note"), "deep")
        extra_keys = {spec.key for spec in loaded_schema if not spec.builtin}
        self.assertIn("site_note", extra_keys)
        site_spec = next(spec for spec in loaded_schema if spec.key == "site_note")
        self.assertEqual(site_spec.label, "Site note")
        self.assertTrue(site_spec.unique)
        self.assertEqual(site_spec.unique_scope, "per_shank")
        saved_site = next(item for item in payload["electrode_attributes"] if item["key"] == "site_note")
        self.assertEqual(saved_site["unique_scope"], "per_shank")
        saved_pot = next(item for item in payload["electrode_attributes"] if item["key"] == "potentiostat_id")
        self.assertEqual(saved_pot["unique_scope"], "per_shank")
        self.assertEqual([p.pad_id for p in loaded_pads], [1, 4])
        self.assertEqual(loaded_pads[0].electrode_eid, 2)
        self.assertEqual(loaded_pads[0].shape, "rect")
        self.assertEqual(loaded_pads[0].height, 4.0)
        self.assertEqual(payload["pads"][0]["pad_id"], 1)
        self.assertEqual(payload["orientation_markers"], [])
        self.assertEqual(document.orientation_markers, [])
        self.assertNotIn("enabled", payload["electrodes"][0])
        self.assertNotIn("enabled", payload["pads"][0])
        self.assertNotIn("pid", payload["pads"][0])
        self.assertNotIn("interface_id", payload["pads"][0])
        self.assertNotIn("system_id", payload["pads"][0])
        self.assertEqual(payload["electrodes"][0]["label_position"], "below")
        self.assertEqual(payload["pads"][0]["label_position"], "below")
        self.assertEqual(payload["electrodes"][0]["label_orientation"], 0)
        self.assertEqual(payload["pads"][0]["label_orientation"], 0)
        self.assertEqual(loaded[0].label_position, "below")
        self.assertEqual(loaded_pads[0].label_position, "below")
        self.assertEqual(loaded[0].label_orientation, 0)
        self.assertEqual(loaded_pads[0].label_orientation, 0)

    def test_legacy_native_without_map_labels_uses_defaults(self) -> None:
        payload = {
            "specification": NATIVE_SPECIFICATION,
            "version": "1.6",
            "si_units": "um",
            "electrodes": [
                {
                    "eid": 1,
                    "x": 0.0,
                    "y": 0.0,
                    "potentiostat_id": 0,
                    "intan_id": "A-000",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "legacy.json")
            Path(path).write_text(json.dumps(payload), encoding="utf-8")
            document = load_array_document(path)
        self.assertEqual(document.map_labels, list(DEFAULT_MAP_LABEL_KEYS))
        self.assertEqual(document.orientation_markers, [])
        self.assertEqual(document.electrodes[0].label_position, "below")
        self.assertEqual(document.electrodes[0].label_orientation, 0)

    def test_legacy_orientation_marker_side_becomes_square(self) -> None:
        payload = {
            "specification": NATIVE_SPECIFICATION,
            "version": "1.10",
            "si_units": "um",
            "electrodes": [
                {
                    "eid": 1,
                    "x": 0.0,
                    "y": 0.0,
                    "potentiostat_id": 0,
                    "intan_id": "A-000",
                }
            ],
            "orientation_markers": [
                {"marker_id": 3, "x": 8.0, "y": 9.0, "side": 24.0},
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "legacy_marker.json")
            Path(path).write_text(json.dumps(payload), encoding="utf-8")
            document = load_array_document(path)
        self.assertEqual(len(document.orientation_markers), 1)
        marker = document.orientation_markers[0]
        self.assertEqual(marker.marker_id, 3)
        self.assertEqual(marker.shape, "square")
        self.assertEqual(marker.radius, 12.0)
        self.assertEqual(marker.height, 0.0)
        electrodes, pads, schema = _sample_array()
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "array.json")
            save_array_to_file(
                path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
                map_labels=[],
            )
            document = load_array_document(path)
        self.assertEqual(document.map_labels, [])

    def test_save_does_not_mutate_missing_extras(self) -> None:
        schema = default_schema() + [
            AttributeSpec(key="site_note", label="Site note", value_type="str", default="")
        ]
        electrode = Electrode(eid=1, x=0.0, y=0.0, extra={})
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "array.json")
            save_array_to_file(path, [electrode], "um", pads=[], electrode_attributes=schema)
            payload = json.loads(Path(path).read_text(encoding="utf-8"))
        self.assertEqual(electrode.extra, {})
        self.assertEqual(payload["electrodes"][0]["attributes"]["site_note"], "")

    def test_legacy_pad_user_ids_are_ignored(self) -> None:
        payload = {
            "specification": NATIVE_SPECIFICATION,
            "version": "1.4",
            "si_units": "um",
            "electrodes": [
                {
                    "eid": 2,
                    "x": 10.0,
                    "y": 20.0,
                    "radius": 12.0,
                    "potentiostat_id": 7,
                    "intan_id": "A-003",
                    "enabled": False,
                }
            ],
            "pads": [
                {
                    "pid": 9,
                    "electrode_eid": 2,
                    "x": 1.0,
                    "y": 2.0,
                    "radius": 10.0,
                    "enabled": False,
                    "interface_id": "P1",
                    "system_id": "INTAN",
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "legacy_pads.json")
            Path(path).write_text(json.dumps(payload), encoding="utf-8")
            _electrodes, pads, _units, _schema = load_array_from_file(path)
        self.assertFalse(hasattr(_electrodes[0], "enabled"))
        self.assertNotIn("enabled", _electrodes[0].extra)
        self.assertEqual(len(pads), 1)
        self.assertEqual(pads[0].pad_id, 9)
        self.assertFalse(hasattr(pads[0], "enabled"))
        self.assertFalse(hasattr(pads[0], "interface_id"))
        self.assertFalse(hasattr(pads[0], "system_id"))

    def test_export_contact_sizes_match_shape_params(self) -> None:
        self.assertEqual(export_contact_sizes("circle", 12.0, 8.0), (12.0, None, None))
        self.assertEqual(export_contact_sizes("square", 6.0, 0.0), (None, 12.0, 12.0))
        self.assertEqual(export_contact_sizes("rect", 12.0, 8.0), (None, 24.0, 16.0))
        self.assertEqual(export_contact_sizes("rect", 12.0, 0.0), (None, 24.0, 24.0))

    def test_spikeinterface_export_roundtrip_restores_native_ids(self) -> None:
        electrodes, pads, schema = _sample_array()
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "probe.json")
            export_spikeinterface_json(
                path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
            )
            payload = build_probeinterface_payload(
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
            )
            probe = payload["probes"][0]
            self.assertEqual(probe["device_channel_indices"], [3, 4])
            self.assertEqual(probe["contact_ids"], ["M3", "M4"])
            self.assertEqual(probe["contact_shapes"], ["rect", "square"])
            self.assertEqual(probe["contact_shape_params"][0]["width"], 24.0)
            self.assertEqual(probe["contact_shape_params"][0]["height"], 16.0)
            self.assertEqual(list(probe["contact_shape_params"][1].keys()), ["width"])
            self.assertEqual(probe["contact_shape_params"][1]["width"], 12.0)
            self.assertEqual(probe["contact_annotations"]["intan_id"], ["A-003", "A-004"])
            self.assertEqual(probe["contact_annotations"]["shank_id"], ["1", ""])
            self.assertEqual(probe["contact_annotations"]["site_note"], ["deep", "shallow"])
            self.assertEqual(probe["contact_annotations"]["pad_id"], [1, 4])
            self.assertEqual(probe["contact_annotations"]["pad_x"], [-30.0, 80.0])
            self.assertEqual(probe["contact_annotations"]["pad_y"], [20.0, 20.0])
            self.assertEqual(probe["contact_annotations"]["pad_shape"], ["rect", "square"])
            self.assertEqual(probe["contact_annotations"]["pad_radius"], [10.0, 10.0])
            self.assertEqual(probe["contact_annotations"]["pad_height"], [4.0, 0.0])
            stored_schema = probe["annotations"][MEA_EDITOR_SCHEMA_ANNOTATION_KEY]
            stored_site = next(item for item in stored_schema if item["key"] == "site_note")
            self.assertEqual(stored_site["label"], "Site note")
            self.assertTrue(stored_site["unique"])
            self.assertEqual(stored_site["unique_scope"], "per_shank")
            self.assertEqual(
                probe["annotations"][MEA_EDITOR_MAP_LABELS_ANNOTATION_KEY],
                list(DEFAULT_MAP_LABEL_KEYS),
            )
            document = load_array_document(path)
        loaded, loaded_pads, units, loaded_schema = (
            document.electrodes,
            document.pads,
            document.si_units,
            document.electrode_attributes,
        )
        self.assertEqual(document.map_labels, list(DEFAULT_MAP_LABEL_KEYS))
        self.assertEqual(units, "um")
        self.assertEqual([pad.pad_id for pad in loaded_pads], [1, 4])
        self.assertEqual(loaded_pads[0].electrode_eid, 2)
        self.assertEqual(loaded_pads[0].x, -30.0)
        self.assertEqual(loaded_pads[0].shape, "rect")
        self.assertEqual(loaded_pads[0].height, 4.0)
        self.assertEqual(loaded_pads[1].shape, "square")
        self.assertEqual(loaded[0].intan_id, "A-003")
        self.assertEqual(loaded[0].manufacturer_id, "M3")
        self.assertEqual(loaded[0].potentiostat_id, 7)
        self.assertEqual(loaded[0].eid, 2)
        self.assertEqual(loaded[0].shape, "rect")
        self.assertEqual(loaded[0].extra.get("site_note"), "deep")
        extra_keys = {spec.key for spec in loaded_schema if not spec.builtin}
        self.assertIn("site_note", extra_keys)
        site_spec = next(spec for spec in loaded_schema if spec.key == "site_note")
        self.assertEqual(site_spec.label, "Site note")
        self.assertTrue(site_spec.unique)
        self.assertEqual(site_spec.unique_scope, "per_shank")
        self.assertEqual(loaded[0].shank_id, "1")

    def test_legacy_probeinterface_without_native_annotations(self) -> None:
        payload = {
            "specification": "probeinterface",
            "version": "0.3.1",
            "probes": [
                {
                    "ndim": 2,
                    "si_units": "um",
                    "annotations": {},
                    "contact_annotations": {},
                    "contact_positions": [[1.0, 2.0]],
                    "contact_plane_axes": [[[1.0, 0.0], [0.0, 1.0]]],
                    "contact_shapes": ["circle"],
                    "contact_shape_params": [{"radius": 12.0}],
                    "device_channel_indices": [3],
                    "contact_ids": ["A-003"],
                    "shank_ids": [""],
                }
            ],
        }
        with tempfile.TemporaryDirectory() as tmp:
            path = str(Path(tmp) / "legacy.json")
            Path(path).write_text(json.dumps(payload), encoding="utf-8")
            loaded, pads, units, _schema = load_array_from_file(path)
        self.assertEqual(units, "um")
        self.assertEqual(pads, [])
        self.assertEqual(loaded[0].potentiostat_id, 3)
        self.assertEqual(loaded[0].intan_id, "A-003")
        self.assertEqual(loaded[0].manufacturer_id, "")

    def test_xlsx_exports_include_pads_and_extras(self) -> None:
        electrodes, pads, schema = _sample_array()
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            analysis_path = str(Path(tmp) / "analysis.xlsx")
            array_path = str(Path(tmp) / "array.xlsx")
            export_analysis_xlsx(analysis_path, electrodes, pads=pads, electrode_attributes=schema)
            export_array_xlsx(
                array_path,
                electrodes,
                pads=pads,
                electrode_attributes=schema,
                si_units="um",
            )
            from openpyxl import load_workbook

            analysis = load_workbook(analysis_path)
            try:
                analysis_sheet = analysis.active
                self.assertEqual(
                    [cell.value for cell in analysis_sheet[1]],
                    [
                        "channel",
                        "row",
                        "col",
                        "shape",
                        "radius",
                        "width",
                        "height",
                        "intan_id",
                        "si_channel",
                        "manufacturer_id",
                        "shank_id",
                        "eid",
                        "site_note",
                        "pad_id",
                        "pad_x",
                        "pad_y",
                        "pad_shape",
                    ],
                )
                self.assertEqual(analysis_sheet["A2"].value, 7)
                self.assertEqual(analysis_sheet["D2"].value, "rect")
                self.assertEqual(analysis_sheet["E2"].value, None)
                self.assertEqual(analysis_sheet["F2"].value, 24.0)
                self.assertEqual(analysis_sheet["G2"].value, 16.0)
                self.assertEqual(analysis_sheet["D3"].value, "square")
                self.assertEqual(analysis_sheet["E3"].value, None)
                self.assertEqual(analysis_sheet["F3"].value, 12.0)
                self.assertEqual(analysis_sheet["G3"].value, 12.0)
                self.assertEqual(analysis_sheet["I2"].value, 3)
                self.assertEqual(analysis_sheet["L2"].value, 2)
                self.assertEqual(analysis_sheet["M2"].value, "deep")
                self.assertEqual(analysis_sheet["N2"].value, 1)
                self.assertEqual(analysis_sheet["O2"].value, -30.0)
                self.assertEqual(analysis_sheet["P2"].value, 20.0)
                self.assertEqual(analysis_sheet["Q2"].value, "rect")
            finally:
                analysis.close()

            workbook = load_workbook(array_path)
            try:
                self.assertEqual(workbook.sheetnames, ["array", "pads", "orientation_markers", "electrode_attributes"])
                self.assertEqual(workbook["orientation_markers"]["A1"].value, "marker_id")
                self.assertIsNone(workbook["orientation_markers"]["A2"].value)
                self.assertEqual(workbook["pads"]["A1"].value, "pad_id")
                self.assertEqual(workbook["pads"]["A2"].value, 1)
                self.assertEqual(workbook["pads"]["E1"].value, "si_channel")
                self.assertEqual(workbook["pads"]["E2"].value, 3)
                self.assertEqual(
                    [cell.value for cell in workbook["array"][1][:12]],
                    [
                        "eid",
                        "potentiostat_id",
                        "intan_id",
                        "si_channel",
                        "manufacturer_id",
                        "row",
                        "col",
                        "shank_id",
                        "shape",
                        "radius",
                        "width",
                        "height",
                    ],
                )
                self.assertEqual(workbook["array"]["D2"].value, 3)
                self.assertEqual(workbook["array"]["I2"].value, "rect")
                self.assertEqual(workbook["array"]["J2"].value, None)
                self.assertEqual(workbook["array"]["K2"].value, 24.0)
                self.assertEqual(workbook["array"]["L2"].value, 16.0)
                self.assertEqual(workbook["array"]["I3"].value, "square")
                self.assertEqual(workbook["array"]["J3"].value, None)
                self.assertEqual(workbook["array"]["K3"].value, 12.0)
                self.assertEqual(workbook["array"]["L3"].value, 12.0)
                self.assertEqual(workbook["array"]["N2"].value, "deep")
                self.assertEqual(workbook["array"]["O2"].value, 1)
                self.assertEqual(
                    [cell.value for cell in workbook["pads"][1][9:13]],
                    ["shape", "radius", "width", "height"],
                )
                self.assertEqual(workbook["pads"]["J2"].value, "rect")
                self.assertEqual(workbook["pads"]["K2"].value, None)
                self.assertEqual(workbook["pads"]["L2"].value, 20.0)
                self.assertEqual(workbook["pads"]["M2"].value, 8.0)
                self.assertEqual(workbook["pads"]["J3"].value, "square")
                self.assertEqual(workbook["pads"]["K3"].value, None)
                self.assertEqual(workbook["pads"]["L3"].value, 20.0)
                self.assertEqual(workbook["pads"]["M3"].value, 20.0)
                self.assertEqual(workbook["pads"]["N2"].value, "deep")
                self.assertEqual(workbook["pads"]["N3"].value, "shallow")
                schema_sheet = workbook["electrode_attributes"]
                self.assertEqual(
                    [cell.value for cell in schema_sheet[1]],
                    ["key", "label", "type", "default", "builtin", "unique", "unique_scope"],
                )
                site_row = next(
                    row for row in schema_sheet.iter_rows(min_row=2, values_only=True) if row[0] == "site_note"
                )
                self.assertEqual(site_row[1], "Site note")
                self.assertEqual(site_row[5], True)
                self.assertEqual(site_row[6], "per_shank")
            finally:
                workbook.close()


    def test_orientation_markers_roundtrip_and_xlsx_only(self) -> None:
        electrodes, pads, schema = _sample_array()
        markers = [
            OrientationMarker(marker_id=2, x=-50.0, y=80.0, radius=15.0, shape="square"),
            OrientationMarker(marker_id=0, x=100.0, y=-10.0, radius=6.0, shape="circle"),
        ]
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            native_path = str(Path(tmp) / "array.json")
            si_path = str(Path(tmp) / "probe.json")
            analysis_path = str(Path(tmp) / "analysis.xlsx")
            array_path = str(Path(tmp) / "array.xlsx")
            save_array_to_file(
                native_path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
                orientation_markers=markers,
            )
            payload = json.loads(Path(native_path).read_text(encoding="utf-8"))
            self.assertEqual(
                [item["marker_id"] for item in payload["orientation_markers"]],
                [0, 2],
            )
            self.assertEqual(payload["orientation_markers"][1]["radius"], 15.0)
            self.assertEqual(payload["orientation_markers"][1]["shape"], "square")
            self.assertNotIn("side", payload["orientation_markers"][1])
            self.assertEqual(payload["orientation_markers"][1]["label_position"], "below")
            self.assertEqual(payload["orientation_markers"][1]["label_orientation"], 0)
            document = load_array_document(native_path)
            self.assertEqual([m.marker_id for m in document.orientation_markers], [0, 2])
            self.assertEqual(document.orientation_markers[1].x, -50.0)
            self.assertEqual(document.orientation_markers[1].radius, 15.0)
            self.assertEqual(document.orientation_markers[1].shape, "square")
            self.assertEqual(document.orientation_markers[0].shape, "circle")

            export_spikeinterface_json(
                si_path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
            )
            si_payload = json.loads(Path(si_path).read_text(encoding="utf-8"))
            dumped = json.dumps(si_payload)
            self.assertNotIn("orientation_marker", dumped)
            self.assertNotIn("marker_id", dumped)
            self.assertNotIn("label_position", dumped)
            self.assertNotIn("label_orientation", dumped)

            export_analysis_xlsx(
                analysis_path,
                electrodes,
                pads=pads,
                electrode_attributes=schema,
                orientation_markers=markers,
            )
            export_array_xlsx(
                array_path,
                electrodes,
                pads=pads,
                electrode_attributes=schema,
                orientation_markers=markers,
            )
            from openpyxl import load_workbook

            analysis = load_workbook(analysis_path)
            try:
                self.assertEqual(analysis.sheetnames, ["array", "orientation_markers"])
                sheet = analysis["orientation_markers"]
                self.assertEqual(
                    [cell.value for cell in sheet[1]],
                    ["marker_id", "x", "y", "shape", "radius", "width", "height"],
                )
                self.assertEqual(sheet["A2"].value, 0)
                self.assertEqual(sheet["B2"].value, 100.0)
                self.assertEqual(sheet["D2"].value, "circle")
                self.assertEqual(sheet["E2"].value, 6.0)
                self.assertEqual(sheet["A3"].value, 2)
                self.assertEqual(sheet["D3"].value, "square")
                self.assertEqual(sheet["F3"].value, 30.0)
                self.assertEqual(sheet["G3"].value, 30.0)
            finally:
                analysis.close()

            workbook = load_workbook(array_path)
            try:
                self.assertEqual(
                    workbook.sheetnames,
                    ["array", "pads", "orientation_markers", "electrode_attributes"],
                )
                self.assertEqual(workbook["orientation_markers"]["A3"].value, 2)
                self.assertEqual(workbook["orientation_markers"]["C3"].value, 80.0)
            finally:
                workbook.close()

    def test_label_position_is_native_only(self) -> None:
        electrodes, pads, schema = _sample_array()
        electrodes[0].label_position = "left"
        electrodes[0].label_orientation = 90
        pads[0].label_position = "right"
        pads[0].label_orientation = 270
        markers = [
            OrientationMarker(
                marker_id=0,
                x=5.0,
                y=6.0,
                radius=8.0,
                height=4.0,
                shape="rect",
                label_position="above",
                label_orientation=180,
            ),
        ]
        with tempfile.TemporaryDirectory(ignore_cleanup_errors=True) as tmp:
            native_path = str(Path(tmp) / "array.json")
            si_path = str(Path(tmp) / "probe.json")
            analysis_path = str(Path(tmp) / "analysis.xlsx")
            array_path = str(Path(tmp) / "array.xlsx")
            save_array_to_file(
                native_path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
                orientation_markers=markers,
            )
            payload = json.loads(Path(native_path).read_text(encoding="utf-8"))
            self.assertEqual(payload["electrodes"][0]["label_position"], "left")
            self.assertEqual(payload["electrodes"][0]["label_orientation"], 90)
            self.assertEqual(payload["pads"][0]["label_position"], "right")
            self.assertEqual(payload["pads"][0]["label_orientation"], 270)
            self.assertEqual(payload["orientation_markers"][0]["label_position"], "above")
            self.assertEqual(payload["orientation_markers"][0]["label_orientation"], 180)
            document = load_array_document(native_path)
            self.assertEqual(document.electrodes[0].label_position, "left")
            self.assertEqual(document.electrodes[0].label_orientation, 90)
            self.assertEqual(document.pads[0].label_position, "right")
            self.assertEqual(document.pads[0].label_orientation, 270)
            self.assertEqual(document.orientation_markers[0].label_position, "above")
            self.assertEqual(document.orientation_markers[0].label_orientation, 180)

            export_spikeinterface_json(
                si_path,
                electrodes,
                "um",
                pads=pads,
                electrode_attributes=schema,
            )
            si_text = Path(si_path).read_text(encoding="utf-8")
            self.assertNotIn("label_position", si_text)
            self.assertNotIn("label_orientation", si_text)

            export_analysis_xlsx(
                analysis_path,
                electrodes,
                pads=pads,
                electrode_attributes=schema,
                orientation_markers=markers,
            )
            export_array_xlsx(
                array_path,
                electrodes,
                pads=pads,
                electrode_attributes=schema,
                orientation_markers=markers,
            )
            from openpyxl import load_workbook

            analysis = load_workbook(analysis_path)
            try:
                analysis_headers = [cell.value for cell in analysis.active[1]]
                marker_headers = [cell.value for cell in analysis["orientation_markers"][1]]
            finally:
                analysis.close()
            workbook = load_workbook(array_path)
            try:
                array_headers = [cell.value for cell in workbook["array"][1]]
                pad_headers = [cell.value for cell in workbook["pads"][1]]
                marker_headers_full = [cell.value for cell in workbook["orientation_markers"][1]]
            finally:
                workbook.close()
        self.assertNotIn("label_position", analysis_headers)
        self.assertNotIn("label_orientation", analysis_headers)
        self.assertNotIn("label_position", marker_headers)
        self.assertNotIn("label_orientation", marker_headers)
        self.assertNotIn("label_position", array_headers)
        self.assertNotIn("label_orientation", array_headers)
        self.assertNotIn("label_position", pad_headers)
        self.assertNotIn("label_orientation", pad_headers)
        self.assertNotIn("label_position", marker_headers_full)
        self.assertNotIn("label_orientation", marker_headers_full)


class VersionTests(unittest.TestCase):
    def test_package_version_is_semver(self) -> None:
        parts = __version__.split(".")
        self.assertGreaterEqual(len(parts), 2)
        self.assertTrue(all(part.isdigit() for part in parts))

    def test_native_format_version(self) -> None:
        self.assertEqual(NATIVE_VERSION, "1.11")

    def test_version_module_matches_package(self) -> None:
        from mea_editor._version import __version__ as raw

        self.assertEqual(raw, __version__)


if __name__ == "__main__":
    unittest.main()
